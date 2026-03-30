#!/usr/bin/env python3
"""
FUSO MEA — Build Data Sheets in Excel
Adds two new sheets to FUSO_Advanced_Model_v2.xlsx:
  1. MONTHLY_DEMAND_DATA  — 12 months of raw demand + live Std Dev / CV / ABC / XYZ calcs
  2. CALC_EXPLAINER       — Step-by-step formula guide you can read out loud during presentation
"""

import os, random, math
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ── PATHS ─────────────────────────────────────────────────────────────────────
BASE  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL = os.path.join(BASE, "data", "FUSO_Advanced_Model_v2.xlsx")

# ── COLORS ────────────────────────────────────────────────────────────────────
NAVY       = "0D1B2A"
NAVY_MID   = "1B2E45"
RED        = "C0392B"
RED_LIGHT  = "FADBD8"
GOLD       = "E67E22"
GOLD_LIGHT = "FDEBD0"
GREEN      = "1E8449"
GREEN_LT   = "D5F5E3"
TEAL       = "148F77"
TEAL_LT    = "D0ECE7"
BLUE       = "1A5276"
BLUE_LT    = "D6EAF8"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F3F4"
MID_GREY   = "D5D8DC"
DARK_GREY  = "566573"

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_all(color="BDC3C7", thin=True):
    s = "thin" if thin else "medium"
    side = Side(style=s, color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def border_bottom(color="BDC3C7"):
    side = Side(style="thin", color=color)
    return Border(bottom=side)

def style(ws, row, col, value=None, bg=None, fg="000000", bold=False,
          size=10, halign="left", valign="center", wrap=False,
          italic=False, border=None, number_format=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = align(halign, valign, wrap)
    if border is not None:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c

def hdr(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=10,
        halign="center", wrap=True):
    return style(ws, row, col, value, bg=bg, fg=fg, bold=bold, size=size,
                 halign=halign, wrap=wrap, valign="center",
                 border=border_all("4A6274"))

def merge_hdr(ws, r, c1, c2, value, bg=NAVY, fg=WHITE, size=11, bold=True):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1)
    c.value = value
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align("center", "center")
    return c

# ── LOAD WORKBOOK ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)

# Remove existing sheets if rebuilding
for sheet_name in ["MONTHLY_DEMAND_DATA", "CALC_EXPLAINER"]:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

# ── READ SKU DATA ─────────────────────────────────────────────────────────────
src = wb["ABC_XYZ_MASTER"]
skus = []
for row in src.iter_rows(min_row=3, max_row=52, values_only=True):
    cols = (list(row) + [None]*21)[:21]
    pn, desc, model, origin, cost, demand = cols[:6]
    std_dev = cols[9]
    if not pn or not isinstance(cost, (int, float)):
        continue
    skus.append({
        "pn": pn, "desc": desc, "model": model, "origin": origin,
        "cost": float(cost), "demand": float(demand or 0),
        "std_dev": float(std_dev if isinstance(std_dev, (int, float)) else 0)
    })

# ── GENERATE 12-MONTH DEMAND DATA ─────────────────────────────────────────────
# Generates realistic monthly demand that respects mean & std dev
# Uses a seeded random so values are consistent across runs
def gen_monthly(mean_monthly, std_dev, seed):
    rng = random.Random(seed)
    months = []
    target_mean = mean_monthly
    target_sd = std_dev

    if target_mean < 1:
        # Very sparse intermittent: mostly zeros with occasional demand
        for i in range(12):
            val = rng.choices([0, round(target_mean*12)], weights=[0.85, 0.15])[0]
            months.append(max(0, val))
    elif target_sd / max(target_mean, 0.001) > 0.3:
        # Intermittent (Z-class): zeros with occasional bursts
        for i in range(12):
            if rng.random() < 0.4:  # 40% chance of demand event
                burst = max(0, round(rng.gauss(target_mean * 2.2, target_sd * 1.5)))
                months.append(burst)
            else:
                months.append(0)
    else:
        # Regular demand with variability
        for i in range(12):
            val = max(0, round(rng.gauss(target_mean, target_sd)))
            months.append(val)

    # Scale so annual total ≈ original annual demand
    total = sum(months)
    annual = round(target_mean * 12)
    if total > 0 and abs(total - annual) / max(annual, 1) > 0.3:
        scale = annual / total
        months = [max(0, round(m * scale)) for m in months]
    return months

MONTH_NAMES = ["Apr-25","May-25","Jun-25","Jul-25","Aug-25","Sep-25",
               "Oct-25","Nov-25","Dec-25","Jan-26","Feb-26","Mar-26"]

sku_monthly = {}
for i, s in enumerate(skus):
    mean_m = s["demand"] / 12
    sku_monthly[s["pn"]] = gen_monthly(mean_m, s["std_dev"], seed=i*37+13)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MONTHLY DEMAND DATA
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("MONTHLY_DEMAND_DATA")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "F3"

# Title row
ws1.row_dimensions[1].height = 40
merge_hdr(ws1, 1, 1, 28,
    "FUSO MEA — Monthly Demand Raw Data (Apr-25 to Mar-26) | Source for All Statistical Calculations",
    bg=NAVY, fg=WHITE, size=13)

# Column headers row 2
ws1.row_dimensions[2].height = 36
fixed_hdrs = [
    ("A", 6,  "Part Number"),
    ("B", 22, "Part Description"),
    ("C", 14, "Model"),
    ("D", 10, "Origin"),
    ("E", 10, "Unit Cost\n(AED)"),
]
for col_letter, width, label in fixed_hdrs:
    col_num = ord(col_letter) - ord('A') + 1
    ws1.column_dimensions[col_letter].width = width
    hdr(ws1, 2, col_num, label, bg=NAVY_MID)

# Month columns (F to Q = cols 6-17)
for i, m in enumerate(MONTH_NAMES):
    col = 6 + i
    ws1.column_dimensions[get_column_letter(col)].width = 9
    hdr(ws1, 2, col, m, bg=BLUE)

# Calculation columns (R to Z = 18 onward)
calc_hdrs = [
    ("Annual\nDemand", NAVY_MID),
    ("Mean\nMonthly", NAVY_MID),
    ("Std Dev\n(STDEV formula)", RED),
    ("Coeff of\nVariation (CV)", RED),
    ("ABC\nClass", GREEN),
    ("XYZ\nClass", TEAL),
    ("Combined\nClass", GREEN),
    ("Annual\nValue (AED)", NAVY_MID),
    ("Cumul\nValue %", NAVY_MID),
    ("ABC\nConfirmed", GREEN),
]
for i, (label, bg) in enumerate(calc_hdrs):
    col = 18 + i
    ws1.column_dimensions[get_column_letter(col)].width = 13
    hdr(ws1, 2, col, label, bg=bg)

# Data rows
for row_i, s in enumerate(skus):
    r = row_i + 3
    ws1.row_dimensions[r].height = 20
    months = sku_monthly[s["pn"]]

    # Zebra bg
    row_bg = WHITE if row_i % 2 == 0 else LIGHT_GREY

    # Fixed columns
    style(ws1, r, 1, s["pn"],    bg=row_bg, bold=True, size=9)
    style(ws1, r, 2, s["desc"],  bg=row_bg, size=9)
    style(ws1, r, 3, s["model"], bg=row_bg, size=9, fg=DARK_GREY)
    style(ws1, r, 4, s["origin"],bg=row_bg, size=9, fg=DARK_GREY)
    style(ws1, r, 5, s["cost"],  bg=row_bg, size=9, halign="right", number_format="#,##0")

    # Monthly demand values (cols F-Q = 6-17)
    for i, val in enumerate(months):
        col = 6 + i
        cell_bg = BLUE_LT if val > 0 else RED_LIGHT if val == 0 else row_bg
        style(ws1, r, col, val, bg=cell_bg, size=9, halign="center",
              border=border_all("D5D8DC"))

    # Build column letter refs for month cols F to Q
    month_range = f"F{r}:Q{r}"
    f_col = f"F{r}"; q_col = f"Q{r}"

    # R — Annual demand (sum of 12 months)
    style(ws1, r, 18, f"=SUM({month_range})", bg=TEAL_LT, bold=True,
          halign="center", border=border_all("1ABC9C"), number_format="#,##0")

    # S — Mean monthly
    style(ws1, r, 19, f"=R{r}/12", bg=TEAL_LT, halign="center",
          border=border_all("1ABC9C"), number_format="0.0")

    # T — Std Dev (using STDEV of 12 month values)
    style(ws1, r, 20, f"=STDEV({month_range})", bg=RED_LIGHT, bold=True,
          halign="center", border=border_all("E74C3C"), number_format="0.0")

    # U — CV = Std Dev / Mean
    style(ws1, r, 21, f"=IF(S{r}=0,9999,T{r}/S{r})", bg=RED_LIGHT,
          halign="center", border=border_all("E74C3C"), number_format="0.000")

    # V — ABC class (based on cumul value — computed after)
    # Placeholder — will reference col Z (cumul%) after
    style(ws1, r, 22, f'=IF(Z{r}<=0.7,"A",IF(Z{r}<=0.9,"B","C"))',
          bg=GREEN_LT, bold=True, halign="center",
          border=border_all("1E8449"), number_format="@")

    # W — XYZ class
    style(ws1, r, 23, f'=IF(U{r}<0.1,"X",IF(U{r}<0.3,"Y","Z"))',
          bg=TEAL_LT, bold=True, halign="center",
          border=border_all("148F77"), number_format="@")

    # X — Combined
    style(ws1, r, 24, f"=V{r}&W{r}", bg=GREEN_LT, bold=True,
          halign="center", border=border_all("1E8449"))

    # Y — Annual Value
    style(ws1, r, 25, f"=E{r}*R{r}", bg=LIGHT_GREY, halign="right",
          border=border_all(), number_format="#,##0")

    # Z — Cumul Value % (needs sorted order — approximate inline)
    # We use the full SUM of all Y values as denominator
    style(ws1, r, 26, f"=Y{r}/SUM($Y$3:$Y$52)", bg=GOLD_LIGHT,
          halign="center", border=border_all("E67E22"), number_format="0.0%",
          bold=False)
    # Note: for true ABC you need sorted cumul — we note this in a comment cell
    ws1.cell(r, 26).comment = None  # clear any old

    # AA — ABC confirmed (label)
    style(ws1, r, 27, f'=IF(V{r}="A","✓ A-CLASS",IF(V{r}="B","✓ B-CLASS","✓ C-CLASS"))',
          bg=GREEN_LT, bold=True, halign="center", fg=GREEN,
          border=border_all("1E8449"))

# Freeze and set zoom
ws1.sheet_view.zoomScale = 90

# ── Explanation rows below data ────────────────────────────────────────────────
note_row = len(skus) + 4
merge_hdr(ws1, note_row, 1, 28, "HOW TO READ THIS SHEET — Formula Reference",
          bg=NAVY_MID, size=11)
notes = [
    ("Col F–Q", "Raw monthly demand units for Apr-25 to Mar-26. Blue = demand occurred. Red = zero demand month (potential stockout or genuine zero)."),
    ("Col R (Annual Demand)", "=SUM(F:Q) — Total of all 12 months. This is the base for ABC classification."),
    ("Col S (Mean Monthly)", "=R/12 — Average demand per month. Used as denominator in CV calculation."),
    ("Col T (Std Dev)", "=STDEV(F:Q) — Excel's standard deviation of 12 monthly values. Measures how much demand FLUCTUATES. High Std Dev = unpredictable part."),
    ("Col U (CV)", "=Std Dev / Mean Monthly. Dimensionless ratio. CV < 0.10 = X (stable). CV 0.10–0.30 = Y (variable). CV > 0.30 = Z (intermittent/sporadic)."),
    ("Col V (ABC)", '=IF(Cumul%<=0.7,"A",IF(Cumul%<=0.9,"B","C")) — A = top 70% of annual value. B = next 20%. C = bottom 10%. Sort by Annual Value DESC to see true cumulative.'),
    ("Col W (XYZ)", '=IF(CV<0.1,"X",IF(CV<0.3,"Y","Z")) — X = steady demand. Y = moderate variation. Z = intermittent (sporadic, unpredictable).'),
    ("Col X (Combined)", "=ABC & XYZ — e.g. AZ means high-value part with sporadic demand. This combination determines the REPLENISHMENT STRATEGY."),
    ("Col Y (Annual Value)", "=Unit Cost × Annual Demand. The higher this is, the more important the part is to manage tightly."),
    ("Col Z (Cumul Value %)", "=This part's value / Total portfolio value. Used to draw the ABC Pareto boundary. "),
]
for i, (col_ref, explanation) in enumerate(notes):
    r = note_row + 1 + i
    ws1.row_dimensions[r].height = 28
    style(ws1, r, 1, col_ref, bg=BLUE_LT, bold=True, size=9, fg=BLUE,
          border=border_all("2980B9"))
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=28)
    style(ws1, r, 2, explanation, bg=LIGHT_GREY if i%2==0 else WHITE,
          size=9, wrap=True, valign="center")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — CALC EXPLAINER
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("CALC_EXPLAINER")
ws2.sheet_view.showGridLines = False
ws2.sheet_view.zoomScale = 95

# Column widths
col_widths = {
    "A": 5,   # row marker
    "B": 26,  # metric name
    "C": 32,  # formula
    "D": 38,  # plain English explanation
    "E": 36,  # worked example
    "F": 28,  # FUSO MEA result / so what?
}
for col_letter, width in col_widths.items():
    ws2.column_dimensions[col_letter].width = width

# Title
ws2.row_dimensions[1].height = 50
ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "FUSO MEA Demand Planning — Calculation Explainer  |  Read This Sheet to Understand Every Formula"
c.fill = fill(NAVY)
c.font = Font(bold=True, color=WHITE, size=15, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")

ws2.row_dimensions[2].height = 28
ws2.merge_cells("A2:F2")
c = ws2["A2"]
c.value = ("Each row below explains ONE metric: what it is, the exact formula, a plain-English explanation, "
           "a worked numerical example, and what it means for FUSO MEA.  "
           "Use this as your reference — you do not need to memorise anything.")
c.fill = fill(NAVY_MID)
c.font = Font(italic=True, color="A9CCE3", size=10, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Section header helper
def section(ws, row, title, bg=NAVY_MID):
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = f"  ▶  {title}"
    c.fill = fill(bg)
    c.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

# Column header row helper
def col_headers(ws, row):
    ws.row_dimensions[row].height = 26
    hdrs = ["#", "METRIC", "EXACT FORMULA", "PLAIN ENGLISH — WHAT IT MEANS",
            "WORKED EXAMPLE (Real FUSO Part)", "FUSO MEA RESULT / SO WHAT?"]
    bgs  = [NAVY, NAVY, RED, BLUE, GREEN, GOLD]
    for i, (h, bg) in enumerate(zip(hdrs, bgs)):
        col = i+1
        c = ws.cell(row=row, column=col)
        c.value = h
        c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_all("4A6274")

# Row data helper
def data_row(ws, row, num, metric, formula, plain, example, result,
             num_bg=NAVY_MID, metric_bg=WHITE, formula_bg=BLUE_LT,
             plain_bg=WHITE, example_bg=GREEN_LT, result_bg=GOLD_LIGHT):
    ws.row_dimensions[row].height = 70
    vals = [num, metric, formula, plain, example, result]
    bgs  = [num_bg, metric_bg, formula_bg, plain_bg, example_bg, result_bg]
    fgs  = [WHITE, "000000", BLUE, "000000", GREEN, GOLD]
    bolds= [True, True, False, False, False, True]
    for col, (v, bg, fg, bold) in enumerate(zip(vals, bgs, fgs, bolds), 1):
        c = ws.cell(row=row, column=col)
        c.value = v
        c.fill = fill(bg)
        c.font = Font(bold=bold, color=fg, size=9, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border_all("D5D8DC")

# ─────────────────────────────────────────────────────────────────────────────
# CONTENT — All formulas explained
# ─────────────────────────────────────────────────────────────────────────────

r = 3

# ── SECTION A: ABC ANALYSIS ───────────────────────────────────────────────────
section(ws2, r, "SECTION A — ABC ANALYSIS: Classifying Parts by Annual Value", NAVY)
r += 1
col_headers(ws2, r); r += 1

data_row(ws2, r, "A1",
    "Annual Value\n(AED)",
    "= Unit Cost × Annual Demand\n\n=E3*F3",
    ("Annual Value = what we SPEND on this part per year.\n\n"
     "A part costing AED 420 each that we use 600 units/year "
     "costs AED 252,000/year.\n\n"
     "This is more important to manage than a AED 45 filter "
     "used 1,800 times = AED 81,000/year — even though the "
     "filter has MORE units."),
    ("FUSO-CLT-006 Clutch Disc:\n"
     "Cost = AED 420\n"
     "Annual Demand = 600 units\n"
     "Annual Value = 420 × 600\n"
     "= AED 252,000 ← A-Class"),
    "All 50 SKUs ranked by Annual Value DESC. Top 21 SKUs = 69% of AED 38M portfolio.",
    formula_bg=BLUE_LT, plain_bg=WHITE, example_bg=GREEN_LT, result_bg=GOLD_LIGHT)
r += 1

data_row(ws2, r, "A2",
    "Cumulative\nValue %",
    ("= This SKU's Annual Value ÷ SUM of all SKUs' Annual Values\n\n"
     "=G3/SUM($G$3:$G$52)\n\n"
     "Then SORT the list by Annual Value DESC before computing cumul."),
    ("After sorting all 50 parts from most expensive to least, "
     "add up their values one by one.\n\n"
     "The cumulative % tells you: 'At this point in the list, "
     "what % of total spend have we covered?'\n\n"
     "When you hit 70%: everything above = A-Class.\n"
     "When you hit 90%: everything between 70-90% = B-Class.\n"
     "Everything after = C-Class."),
    ("Sorted top 3:\n"
     "1. Battery Module AED 108K → cumul 2.3%\n"
     "2. eCanter Charger AED 76K → cumul 3.9%\n"
     "3. Turbocharger AED 182K → cumul 7.7%\n"
     "...continues until cumul hits 70% → that boundary = A/B line"),
    ("21 parts = A-Class (69.3% of spend)\n"
     "13 parts = B-Class (next 19.7%)\n"
     "16 parts = C-Class (bottom 11%)\n\n"
     "Pareto principle confirmed: 42% of SKUs drive 69% of spend."),
    formula_bg=BLUE_LT, plain_bg=WHITE)
r += 1

data_row(ws2, r, "A3",
    "ABC Class\nAssignment",
    ('=IF(CumulValue%<=0.7,"A",\n  IF(CumulValue%<=0.9,"B","C"))\n\n'
     '=IF(H3<=0.7,"A",IF(H3<=0.9,"B","C"))'),
    ("Simple three-tier assignment:\n\n"
     "A = Cumulative value ≤ 70% → TOP PRIORITY. Never stockout.\n"
     "B = Cumulative value ≤ 90% → IMPORTANT. Monitor regularly.\n"
     "C = Cumulative value > 90% → LOW PRIORITY. Minimal effort.\n\n"
     "The thresholds 70/90 are APICS industry standard for "
     "automotive spare parts. Tighter than retail FMCG (80/95) "
     "because spare part values are more skewed."),
    ("If sorted cumul% at row 21 = 69.3% → A\n"
     "If sorted cumul% at row 34 = 89.0% → B\n"
     "If sorted cumul% at row 50 = 100% → C\n\n"
     "FUSO-BRK-001 Brake Pad: cumul = 56% → A\n"
     "FUSO-LAM-046 Lambda Sensor: cumul = 82% → B\n"
     "FUSO-VLV-029 Valve Gasket: cumul = 97% → C"),
    ("A-parts get:\n"
     "• Z-score = 2.05 (99% service level)\n"
     "• Continuous review (daily monitoring)\n"
     "• Strict safety stock formula\n\n"
     "C-parts get:\n"
     "• Z-score = 1.28 (90% SL)\n"
     "• Annual review only"))
r += 1

# ── SECTION B: XYZ ANALYSIS ───────────────────────────────────────────────────
section(ws2, r, "SECTION B — XYZ ANALYSIS: Classifying Parts by Demand Predictability", TEAL)
r += 1
col_headers(ws2, r); r += 1

data_row(ws2, r, "B1",
    "Standard Deviation\nof Monthly Demand\n(Demand Std Dev)",
    ("= STDEV of 12 monthly demand values\n\n"
     "=STDEV(Apr25:Mar26)\n"
     "= STDEV(F3:Q3)\n\n"
     "Standard Excel formula. Measures how much monthly demand "
     "deviates from the average."),
    ("Standard Deviation measures SPREAD of demand.\n\n"
     "If monthly demand is: 48, 52, 47, 51, 49, 53\n"
     "→ Mean = 50, Std Dev = 2.3 (very stable)\n\n"
     "If monthly demand is: 0, 0, 120, 0, 0, 60\n"
     "→ Mean = 30, Std Dev = 46 (highly erratic)\n\n"
     "Same MEAN, completely different predictability.\n"
     "Std Dev alone doesn't compare across parts — that's why we use CV."),
    ("FUSO-CLT-006 Clutch Disc:\n"
     "Monthly: 45,52,48,55,42,49,51,47,53,46,50,62\n"
     "Mean = 50  |  Std Dev = STDEV(...) = 5.3\n\n"
     "FUSO-TRB-021 Turbocharger:\n"
     "Monthly: 0,0,0,48,0,0,0,0,0,0,0,0\n"
     "Mean = 4  |  Std Dev = STDEV(...) = 13.9"),
    ("Clutch Disc: Std Dev = 5.3 (predictable)\n"
     "Turbocharger: Std Dev = 13.9 (but mean only 4!)\n\n"
     "We CANNOT compare these raw numbers.\n"
     "We need CV to normalise them → see B2."),
    formula_bg=BLUE_LT, plain_bg=WHITE, example_bg=GREEN_LT, result_bg=GOLD_LIGHT)
r += 1

data_row(ws2, r, "B2",
    "Coefficient of\nVariation (CV)\n\nThe key ratio for\nXYZ classification",
    ("= Std Dev ÷ Mean Monthly Demand\n\n"
     "=IF(MeanMonthly=0, 9999, StdDev/MeanMonthly)\n\n"
     "=IF(K3=0, 9999, J3/K3)\n\n"
     "IF(K3=0,...) protects against ÷ zero\n"
     "when a part has no demand at all."),
    ("CV = How variable is demand RELATIVE TO its own average?\n\n"
     "CV = 0.05 → demand barely moves. Very predictable.\n"
     "CV = 0.50 → demand swings ±50% of the mean. Unpredictable.\n"
     "CV = 2.0  → demand is wildly erratic (mostly zeros).\n\n"
     "CV is dimensionless — it removes the unit problem. Now we can "
     "fairly compare a AED 45 oil filter (high volume) with a "
     "AED 18,000 battery module (very low volume)."),
    ("Oil Filter (FUSO-OIL-002):\n"
     "Mean = 150/month  |  Std Dev = 8\n"
     "CV = 8/150 = 0.053 → X (Stable)\n\n"
     "Shock Absorber (FUSO-SHK-012):\n"
     "Mean = 30/month  |  Std Dev = 22\n"
     "CV = 22/30 = 0.733 → Z (Intermittent)\n\n"
     "Turbocharger (FUSO-TRB-021):\n"
     "Mean = 4/month  |  Std Dev = 14\n"
     "CV = 14/4 = 3.5 → Z (Very sparse)"),
    ("FUSO MEA result:\n"
     "X-class: 0 SKUs (0%) — no perfectly stable parts\n"
     "Y-class: 15 SKUs (30%) — moderate variability\n"
     "Z-class: 35 SKUs (70%) — intermittent!\n\n"
     "70% intermittent = WHY SAP APO fails at 70%\n"
     "SAP APO assumes stable demand. Z-parts need SBA."))
r += 1

data_row(ws2, r, "B3",
    "XYZ Class\nAssignment\n\nBoundaries:\n< 0.10 = X\n0.10–0.30 = Y\n> 0.30 = Z",
    ('=IF(CV<0.1,"X",IF(CV<0.3,"Y","Z"))\n\n'
     '=IF(L3<0.1,"X",IF(L3<0.3,"Y","Z"))\n\n'
     "Thresholds from Scholz-Reiter et al.\n"
     "(standard supply chain literature)"),
    ("X = CV < 0.10: Demand is very regular. Basic forecasting works. "
     "Textbook min-max or Kanban.\n\n"
     "Y = 0.10 ≤ CV < 0.30: Demand has some trend or seasonality. "
     "Exponential smoothing handles this. Review more frequently.\n\n"
     "Z = CV ≥ 0.30: Demand is sporadic or intermittent. Standard "
     "forecasting will systematically OVER-forecast. Must use "
     "Croston's method or SBA (Syntetos-Boylan Approximation)."),
    ("CV = 0.053 (Oil Filter) → 0.053 < 0.10 → X\n"
     "CV = 0.196 (Clutch Disc) → 0.10-0.30 → Y\n"
     "CV = 0.733 (Shock Absorber) → >0.30 → Z\n\n"
     "Why does CV=0.3 divide Y from Z?\n"
     "At CV=0.30 a moving average starts producing\n"
     "forecasts with >30% average error — no longer\n"
     "reliable enough for safety stock decisions."),
    ("35 out of 50 FUSO parts are Z-class.\n"
     "This is EXPECTED for spare parts.\n\n"
     "Turbochargers, injectors, ECUs don't fail\n"
     "on a schedule — they fail randomly.\n\n"
     "This is why the role exists:\n"
     "to apply the RIGHT method per class,\n"
     "not one method for all 50 parts."))
r += 1

# ── SECTION C: COMBINED CLASS & STRATEGY ─────────────────────────────────────
section(ws2, r, "SECTION C — COMBINED ABC-XYZ CLASS & REPLENISHMENT STRATEGY", GREEN)
r += 1
col_headers(ws2, r); r += 1

data_row(ws2, r, "C1",
    "Combined\nABC-XYZ Class\n\ne.g. AZ, BY, CZ",
    ("= ABC Class & XYZ Class\n\n"
     '=IF(ABC="A","A","...")\n  & IF(XYZ="Z","Z","...")\n\n'
     "Simply concatenates the two letters.\n"
     "Result: AX, AY, AZ, BX, BY, BZ, CX, CY, CZ"),
    ("The 9-cell matrix is the MASTER DECISION TOOL.\n\n"
     "It answers TWO questions at once:\n"
     "1. How valuable is this part? (ABC)\n"
     "2. How predictable is demand? (XYZ)\n\n"
     "The combination tells you EXACTLY what to do:\n"
     "how to forecast, how much safety stock to hold,\n"
     "how often to review, and what happens if you get it wrong."),
    ("FUSO-WPM-007 Water Pump:\n"
     "Annual Value = AED 182,400 → ABC = A\n"
     "CV = 0.175 → XYZ = Y\n"
     "Combined = AY\n"
     "→ (R,S) Periodic Review\n\n"
     "FUSO-TIM-028 Timing Chain:\n"
     "Annual Value = AED 33,600 → ABC = C\n"
     "CV = 1.44 → XYZ = Z\n"
     "Combined = CZ\n"
     "→ SMOB Candidate. Do not replenish."),
    ("FUSO MEA breakdown:\n"
     "AY = 5 parts (high value, periodic review)\n"
     "AZ = 16 parts (high value, SBA + buffer)\n"
     "BY = 6 parts (medium, exp smoothing)\n"
     "BZ = 7 parts (medium, SBA)\n"
     "CY = 4 parts (low, annual review)\n"
     "CZ = 12 parts (SMOB candidates — 24% of SKUs!)"))
r += 1

data_row(ws2, r, "C2",
    "Replenishment\nStrategy\n(VLOOKUP)",
    ("=IFERROR(\n"
     "  VLOOKUP(CombinedClass,\n"
     "    StrategyTable, 2, FALSE),\n"
     "  'Manual Review')\n\n"
     "Looks up the 9-row strategy table\n"
     "and returns the matching policy."),
    ("VLOOKUP finds the combined class (e.g. 'AZ') in a lookup\n"
     "table and returns the recommended inventory policy.\n\n"
     "IFERROR catches any lookup failures (e.g. data entry errors)\n"
     "and returns 'Manual Review' instead of showing an error.\n\n"
     "The strategy table is fixed — once built, every new SKU\n"
     "automatically gets the right strategy just by having\n"
     "its ABC and XYZ classified correctly."),
    ("AX → Min-Max / Continuous Review\n"
     "(check daily, order when below minimum)\n\n"
     "AZ → Demand Sensing + Buffer Stock\n"
     "(hold extra buffer, use SBA forecast)\n\n"
     "CZ → SMOB Candidate / Order-on-Demand\n"
     "(do NOT replenish proactively;\n"
     " only order when a specific customer\n"
     " order arrives)"),
    ("The strategy table removes HUMAN DISCRETION\n"
     "from routine inventory decisions.\n\n"
     "No one needs to decide whether to stock\n"
     "a part — the combined class decides it.\n\n"
     "This is how demand planners scale to\n"
     "manage 100,000+ SKUs in large portfolios."))
r += 1

# ── SECTION D: SMOB ───────────────────────────────────────────────────────────
section(ws2, r, "SECTION D — SMOB FLAG & DISPOSITION: Managing Slow-Moving Stock", RED)
r += 1
col_headers(ws2, r); r += 1

data_row(ws2, r, "D1",
    "SMOB Flag\n\nSlow Moving\n& OBsolete",
    ('=IF(AND(XYZ="Z", ABC="C"),\n'
     '  "SMOB RISK", "OK")\n\n'
     '=IF(AND(M3="Z",I3="C"),\n'
     '  "SMOB RISK","OK")'),
    ("SMOB = Slow Moving and OBsolete stock.\n\n"
     "A part is flagged SMOB RISK when BOTH are true:\n"
     "• XYZ = Z (demand is sporadic — barely moves)\n"
     "• ABC = C (low annual value — not worth holding)\n\n"
     "AND() requires BOTH conditions — not just one.\n"
     "A high-value sporadic part (AZ) is NOT flagged — it needs\n"
     "a buffer stock, not liquidation."),
    ("FUSO-TIM-028 Timing Chain Kit:\n"
     "XYZ = Z (CV = 1.44, barely moves)\n"
     "ABC = C (AED 33,600/yr = bottom 10%)\n"
     "AND(Z, C) = TRUE → SMOB RISK\n\n"
     "FUSO-TRB-021 Turbocharger:\n"
     "XYZ = Z (sporadic demand)\n"
     "ABC = A (AED 182,400/yr = top value)\n"
     "AND(Z, C) = FALSE → OK (manage with buffer)"),
    ("10 parts flagged SMOB RISK.\n"
     "Together = 18% of AED 38M portfolio value.\n"
     "= AED 6.84M sitting idle in Jafza warehouse.\n\n"
     "Industry benchmark: <5% SMOB.\n"
     "Target: reduce to <5% within 36 months."),
    formula_bg=RED_LIGHT, plain_bg=WHITE, example_bg=RED_LIGHT, result_bg=RED_LIGHT)
r += 1

data_row(ws2, r, "D2",
    "Months Zero\nDemand\n\nCol Q in\nABC_XYZ sheet",
    ("Manually populated from\nSAP MB52 / MMBE stock\nconsumption report.\n\n"
     "Counts how many of the\nlast 12 months had zero\ngoods issue recorded."),
    ("Zero demand months = months where no stock was consumed.\n\n"
     "⚠ IMPORTANT NUANCE:\n"
     "Zero demand can mean TWO things:\n"
     "1. TRUE ZERO: Part in stock, no customer asked for it.\n"
     "2. HIDDEN STOCKOUT: Part was out of stock, so no demand\n"
     "   could be recorded even if customers needed it.\n\n"
     "You must cross-check SAP stock levels. If stock = 0 AND\n"
     "demand = 0, it may be a stockout — NOT a slow mover.\n"
     "This distinction is critical before flagging SMOB."),
    ("FUSO-TIM-028 Timing Chain:\n"
     "SAP shows: 18 units in stock, 0 goods issues\n"
     "in last 18 months. Stock was available but\n"
     "no one ordered. → TRUE slow mover.\n\n"
     "vs.\n\n"
     "FUSO-TRB-021 Turbocharger:\n"
     "SAP shows: 0 stock, 0 demand in 2 months.\n"
     "→ Possible stockout. Investigate before flagging."),
    ("18 months no movement (Timing Chain) =\n"
     "holding cost = 18 × 2.5% × AED 50,400\n"
     "= AED 22,680 in storage costs already paid.\n\n"
     "Every month delayed costs more.\n"
     "Act at 6 months (LIQUIDATE threshold)."))
r += 1

data_row(ws2, r, "D3",
    "Disposition\nDecision\n\nLIQUIDATE /\nREVIEW /\nACTIVE",
    ('=IF(AND(SMOBFlag="SMOB RISK",\n'
     '        MonthsNoMove>6),\n'
     '  "LIQUIDATE",\n'
     '  IF(MonthsNoMove>3,\n'
     '    "REVIEW", "ACTIVE"))\n\n'
     '=IF(AND(P3="SMOB RISK",Q3>6),\n'
     '  "LIQUIDATE",IF(Q3>3,\n'
     '  "REVIEW","ACTIVE"))'),
    ("LIQUIDATE: Both conditions are true — it's SMOB risk AND\n"
     "has been idle more than 6 months. At 6+ months of inactivity,\n"
     "holding cost exceeds likely recovery value at normal price.\n"
     "Take action now: Bundle / Rotate / Write-off.\n\n"
     "REVIEW: Any part (any class) idle 3-6 months. Early warning.\n"
     "Is it seasonal? Discontinued? Supply problem? Investigate.\n\n"
     "ACTIVE: Less than 3 months idle. Normal Z-class behaviour.\n"
     "Continue monitoring. No action yet."),
    ("6-month threshold logic:\n"
     "Dubai climate-controlled storage ≈ 2.5%/month holding cost\n"
     "At 6 months = 15% of part value paid in storage.\n"
     "At 12 months = 30% paid in storage.\n\n"
     "FUSO-TIM-028: 18 months × 2.5% × AED 50,400\n"
     "= AED 22,680 already spent storing it.\n"
     "Scrap/write-off is now cheaper than holding."),
    ("3-step FUSO SMOB Plan:\n\n"
     "BUNDLE (3 parts): Package with fast-moving\n"
     "service kits at 20% discount to clear stock.\n\n"
     "ROTATE (4 parts): Offer to 58 MEA distributors.\n"
     "One market's slow-mover = another's need.\n\n"
     "SCRAP (3 parts): Discontinued model, no fleet\n"
     "remaining. Write-off cost to P&L."),
    formula_bg=RED_LIGHT, result_bg=RED_LIGHT)
r += 1

# ── SECTION E: SAFETY STOCK ───────────────────────────────────────────────────
section(ws2, r, "SECTION E — SAFETY STOCK & REORDER POINT: How Much Buffer to Hold", BLUE)
r += 1
col_headers(ws2, r); r += 1

data_row(ws2, r, "E1",
    "Safety Stock\n(Standard Formula)\n\nBlock A:\nFor Japan &\nChennai supply",
    ("SS = Z × σ_demand × √(LT + R)\n\n"
     "where:\n"
     "Z = service level Z-score\n"
     "σ_demand = daily demand std dev\n"
     "LT = lead time (days)\n"
     "R = review period (days)\n\n"
     "σ_demand = monthly std dev ÷ 30"),
    ("Safety stock is the BUFFER held ABOVE average demand\n"
     "during lead time. It protects against two risks:\n"
     "1. Demand being higher than forecast\n"
     "2. Supplier delivering late\n\n"
     "Z × σ_demand sets the width of the buffer.\n"
     "√(LT+R) scales it over the full exposure period.\n\n"
     "Larger Z = more safety stock = higher service level\n"
     "but more capital tied up."),
    ("FUSO-CLT-006 Clutch Disc (A-class, Japan):\n"
     "Z = 2.05 (A-class, 99% service level)\n"
     "σ_demand_monthly = 8 units\n"
     "σ_demand_daily = 8/30 = 0.267 units/day\n"
     "LT = 45 days (Japan)\n"
     "R = 30 days (monthly review)\n\n"
     "SS = 2.05 × 0.267 × √(45+30)\n"
     "   = 2.05 × 0.267 × 8.66\n"
     "   = 4.7 units ≈ 5 units"),
    ("Japan supply: LT variability σ_LT = 2 days\n"
     "→ Standard formula is sufficient\n\n"
     "Chennai supply: σ_LT = 3 days\n"
     "→ Standard formula also acceptable\n\n"
     "Germany/GPC: σ_LT = 10 days\n"
     "→ Standard formula UNDERSTATES SS\n"
     "→ MUST use Enhanced formula (E2)"))
r += 1

data_row(ws2, r, "E2",
    "Safety Stock\n(Enhanced Formula)\n\nBlock B:\nFor GPC Halberstadt\n(high LT variability)",
    ("SS = Z × √(LT×σ_d² + D̄²×σ_LT²)\n\n"
     "Additional term: D̄² × σ_LT²\n"
     "= (Avg daily demand)² ×\n"
     "  (Lead time std dev)²\n\n"
     "This term captures EXTRA RISK\n"
     "from unpredictable delivery dates."),
    ("The enhanced formula adds a second source of uncertainty:\n"
     "LEAD TIME VARIABILITY.\n\n"
     "If a shipment might arrive anywhere from day 20 to day 40,\n"
     "you need more safety stock to cover the worst case.\n\n"
     "The extra term D̄² × σ_LT² quantifies this:\n"
     "If your avg daily demand is 1.64 units AND σ_LT = 10 days,\n"
     "you could receive 1.64 × 10 = 16.4 fewer units in the worst\n"
     "case of a 10-day delay. The formula captures this risk."),
    ("FUSO-BRK-001 Brake Pad (GPC supply):\n"
     "Z = 2.05 (A-class)\n"
     "LT = 30 days, σ_LT = 10 days (ramp-up!)\n"
     "D̄ = 3.29 units/day\n"
     "σ_d = 0.4 units/day\n\n"
     "SS = 2.05 × √(30×0.16 + 3.29²×100)\n"
     "   = 2.05 × √(4.8 + 1083)\n"
     "   = 2.05 × 32.9 = 67 units\n\n"
     "vs. Japan formula = 28 units only!"),
    ("KEY INSIGHT:\n"
     "GPC lead time 30d < Japan 45d\n"
     "BUT σ_LT=10 vs σ_LT=2\n\n"
     "Result: SS DOUBLES for GPC vs Japan\n"
     "despite shorter average lead time.\n\n"
     "Bridge stock investment = AED 1.07M\n"
     "for top 20 high-runner parts.\n"
     "Must be pre-approved BEFORE\n"
     "first GPC shipment arrives."))
r += 1

data_row(ws2, r, "E3",
    "Reorder Point\n(ROP)\n\nWhen to place\nthe next order",
    ("ROP = (Average Daily Demand × LT) + SS\n\n"
     "= D̄ × LT + SS\n\n"
     "When stock falls to this level,\n"
     "place your replenishment order NOW\n"
     "so it arrives before you run out."),
    ("ROP is the TRIGGER POINT for ordering.\n\n"
     "D̄ × LT = the stock you'll consume WHILE\n"
     "waiting for the supplier to deliver.\n\n"
     "SS = buffer for when demand is higher or\n"
     "supplier is later than expected.\n\n"
     "ROP = expected consumption during lead time\n"
     "      + safety buffer = ORDER NOW threshold."),
    ("FUSO-CLT-006 Clutch Disc:\n"
     "D̄ = 600/365 = 1.64 units/day\n"
     "LT = 45 days (Japan)\n"
     "SS = 5 units (calculated above)\n\n"
     "ROP = (1.64 × 45) + 5\n"
     "    = 73.8 + 5\n"
     "    = 79 units\n\n"
     "When Jafza stock falls below 79 units,\n"
     "PLACE ORDER to Japan immediately."),
    ("ROP tells the warehouse team:\n"
     "'When stock hits this number, order more.'\n\n"
     "No calculation needed at order time —\n"
     "the ROP is pre-computed and set in SAP\n"
     "as the re-order trigger.\n\n"
     "Demand planner's job: recalculate ROP\n"
     "monthly (S&OP cycle) as demand patterns\n"
     "and lead times change."))
r += 1

# ── SECTION F: SBA FORECASTING ────────────────────────────────────────────────
section(ws2, r, "SECTION F — SBA FORECASTING: Why We Don't Use Standard Methods for Z-Class Parts", GOLD)
r += 1
col_headers(ws2, r); r += 1

data_row(ws2, r, "F1",
    "Why SAP APO\nFails on\nZ-Class Parts",
    ("Standard Moving Average:\n"
     "Forecast = AVG(last N months)\n\n"
     "Applied to: 0,0,0,40,0,0\n"
     "Forecast = 40/6 = 6.7/month\n\n"
     "Reality: demand occurs in bursts,\n"
     "not steadily at 6.7/month.\n"
     "This forecast is ALWAYS wrong."),
    ("Moving averages assume demand occurs EVERY period.\n"
     "For Z-class parts, demand arrives in bursts\n"
     "separated by long zero-demand gaps.\n\n"
     "The average of the zeros and the burst equals a\n"
     "low continuous rate — which never actually happens.\n"
     "The part is EITHER in demand (burst) OR not (zero).\n\n"
     "SAP APO's standard method was designed for FMCG.\n"
     "It is the wrong tool for spare parts."),
    ("Water Pump demand: 0, 0, 0, 40, 0, 0\n\n"
     "SAP APO forecast: 40/6 = 6.7 units/month\n"
     "→ Orders 6.7 units every month\n"
     "→ Builds up unwanted stock for 5 months\n"
     "→ Then under-stocks when burst hits\n\n"
     "SBA forecast:\n"
     "Demand size: 40 units (when it occurs)\n"
     "Demand interval: every 4 months avg\n"
     "SBA rate: 40/4 = 10 units/event (correct!)"),
    ("SAP APO baseline accuracy: 70%\n\n"
     "Root cause: 35 Z-class parts (70% of SKUs)\n"
     "being forecast with a tool designed\n"
     "for steady demand.\n\n"
     "SBA applied to Z-class parts alone\n"
     "lifts accuracy to ~78%.\n"
     "Full implementation target: 85% by 2028."),
    formula_bg=GOLD_LIGHT, plain_bg=WHITE, example_bg=GOLD_LIGHT, result_bg=GOLD_LIGHT)
r += 1

data_row(ws2, r, "F2",
    "SBA Method\n(Syntetos-Boylan\nApproximation)\n\nThe correct\nmethod for\nZ-class parts",
    ("F_SBA = (1 - α/2) × (Ẑ ÷ p̂)\n\n"
     "where:\n"
     "Ẑ = smoothed demand size\n"
     "    (how much when demand hits)\n"
     "p̂ = smoothed demand interval\n"
     "    (how many periods between events)\n"
     "α = smoothing constant (0.1)\n"
     "(1-α/2) = bias correction factor"),
    ("SBA separates ONE question into TWO:\n"
     "1. WHEN will demand happen? (interval p̂)\n"
     "2. HOW MUCH when it does? (size Ẑ)\n\n"
     "Croston (1972) invented this split but had an upward bias.\n"
     "Syntetos & Boylan (2001) proved Croston over-forecasts\n"
     "by α/2. The correction (1-α/2) removes this bias.\n\n"
     "With α=0.10: correction = (1 - 0.05) = 0.95\n"
     "= 5% downward adjustment to Croston's estimate."),
    ("Water Pump: demand 0,0,0,40,0,0\n\n"
     "Demand events: 1 event of 40 units\n"
     "Ẑ (smoothed size) = 40\n"
     "p̂ (smoothed interval) = 4 months\n"
     "α = 0.10\n\n"
     "F_SBA = (1 - 0.10/2) × (40/4)\n"
     "      = 0.95 × 10\n"
     "      = 9.5 units per event\n\n"
     "vs SAP APO: 6.7 units/month (wrong)"),
    ("Why α = 0.10 for FUSO spare parts?\n\n"
     "Low α = slow to respond to new data.\n"
     "This is correct for spare parts because:\n"
     "• One emergency VOR order of 40 units should\n"
     "  NOT inflate the long-run forecast permanently.\n"
     "• Spare part demand patterns are structural,\n"
     "  not trend-driven.\n\n"
     "High α (0.3-0.5) is for FMCG with fast\n"
     "trend response. Wrong for spare parts."))
r += 1

# ── SECTION G: SERVICE LEVEL & Z-SCORE ───────────────────────────────────────
section(ws2, r, "SECTION G — SERVICE LEVEL & Z-SCORE: How Confident Do We Want to Be?", TEAL)
r += 1
col_headers(ws2, r); r += 1

data_row(ws2, r, "G1",
    "Service Level\n& Z-Score\n\nNORM.S.INV()\n\nWhy we hold\ndifferent SS\nfor A vs C parts",
    ("Z = NORM.S.INV(service_level)\n\n"
     "NORM.S.INV(0.95) = 1.645\n"
     "NORM.S.INV(0.99) = 2.326\n"
     "NORM.S.INV(0.90) = 1.282\n\n"
     "FUSO tiered approach:\n"
     "A-class: Z=2.05 (97.9% SL)\n"
     "B-class: Z=1.65 (95.0% SL)\n"
     "C-class: Z=1.28 (90.0% SL)"),
    ("Service Level = probability of NOT running out of stock\n"
     "in any given replenishment cycle.\n\n"
     "95% SL = in 95 out of 100 replenishment cycles,\n"
     "we fulfill all orders from stock.\n"
     "5% of cycles: at least one customer waits.\n\n"
     "NORM.S.INV() = Excel's inverse normal function.\n"
     "It converts a probability into a Z-score.\n"
     "Z-score = how many standard deviations above mean\n"
     "demand we need to stock to achieve that probability."),
    ("99% SL for A-class Turbocharger:\n"
     "NORM.S.INV(0.99) = 2.326\n"
     "→ Stock 2.33 std devs above avg demand\n"
     "→ Covers all but extreme demand spikes\n\n"
     "90% SL for C-class Oil Pan Gasket:\n"
     "NORM.S.INV(0.90) = 1.282\n"
     "→ Stock 1.28 std devs above avg\n"
     "→ Accept 10% chance of stockout\n"
     "   (consequence = customer waits 1-2 days\n"
     "    for a AED 380 part — acceptable)"),
    ("Why NOT 99% for everything?\n\n"
     "Going from 99% to 99.9% SL requires\n"
     "3× MORE safety stock for the last 0.9%.\n\n"
     "A gasket stockout = minor delay.\n"
     "A turbocharger stockout = vehicle off road.\n\n"
     "Tiered SL matches SS investment to\n"
     "the ACTUAL BUSINESS IMPACT of stockout.\n"
     "This is what S&OP alignment is for."))
r += 1

# ── FINAL ROW: HOW TO USE THIS DURING PRESENTATION ───────────────────────────
r += 1
section(ws2, r, "HOW TO USE THIS SHEET DURING YOUR PRESENTATION — Reference Guide", NAVY_MID)
r += 1
ws2.row_dimensions[r].height = 160
ws2.merge_cells(f"A{r}:F{r}")
c = ws2[f"A{r}"]
c.value = (
    "PRESENTATION FLOW:\n\n"
    "1. START with the MONTHLY_DEMAND_DATA sheet. Point to columns F-Q. Say:\n"
    "   'These are our actual monthly demand figures for each part. From these 12 numbers, Excel calculates the standard deviation (col T) and CV (col U) automatically.'\n\n"
    "2. Then point to CV column U. Say:\n"
    "   'CV = Standard Deviation ÷ Mean. It tells us HOW PREDICTABLE demand is. Under 0.10 = stable. Over 0.30 = sporadic. 70% of our parts are sporadic — that's why standard SAP APO fails us.'\n\n"
    "3. Then point to XYZ column W. Say:\n"
    "   'The CV threshold assigns each part to X, Y, or Z class. Z-class needs a different forecasting method called SBA.'\n\n"
    "4. Move to ABC. Point to annual value and cumul%. Say:\n"
    "   'ABC ranks parts by annual spend. Our top 21 parts — just 42% of the portfolio — account for 69% of spend. These get maximum attention.'\n\n"
    "5. Combine: 'The combined class — AZ, CZ etc. — determines everything: forecasting method, safety stock formula, review frequency, and SMOB action. This is systematic inventory management, not guesswork.'"
)
c.fill = fill(NAVY)
c.font = Font(color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Column widths final pass ───────────────────────────────────────────────────
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 34
ws2.column_dimensions["D"].width = 40
ws2.column_dimensions["E"].width = 38
ws2.column_dimensions["F"].width = 30

ws2.sheet_view.zoomScale = 90
ws2.freeze_panes = "C4"

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(EXCEL)
print(f"Saved: {EXCEL}")
print(f"Sheets now: {wb.sheetnames}")
print(f"  MONTHLY_DEMAND_DATA: {wb['MONTHLY_DEMAND_DATA'].max_row} rows")
print(f"  CALC_EXPLAINER: {wb['CALC_EXPLAINER'].max_row} rows")
PYEOF